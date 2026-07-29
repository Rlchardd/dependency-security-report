# Importa datetime e timezone.
# datetime representa data e horário.
# timezone será usado para remover corretamente
# o fuso horário antes de escrever a data no Excel.
from datetime import datetime, timezone

# Importa Path para representar o caminho
# onde a planilha será salva.
from pathlib import Path

# Importa Workbook, responsável por criar
# um novo arquivo Excel.
from openpyxl import Workbook

# Importa classes usadas para estilizar células.
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)

# Importa uma função que transforma o número
# de uma coluna em sua letra correspondente.
#
# Exemplo:
# 1 → A
# 2 → B
# 3 → C
from openpyxl.utils import get_column_letter

# Importa o tipo Worksheet para documentar
# qual tipo de objeto representa a aba da planilha.
from openpyxl.worksheet.worksheet import Worksheet

# Importa o model consolidado que será
# utilizado para preencher cada linha.
from src.domain.models import PackageReportModel

# Importa a exceção específica da planilha.
from src.share.exceptions import SpreadsheetError


# Classe responsável por criar e salvar
# o relatório em formato Excel.
class ExcelRepository:

    # Define os títulos das colunas.
    #
    # Como o valor não deverá mudar durante a execução,
    # usamos letras maiúsculas.
    HEADERS: tuple[str, ...] = (
        "Dependência",
        "Versão no projeto",
        "Última versão",
        "Descrição",
        "Score",
        "Vulnerabilidades",
        "Licença",
        "Última publicação",
    )

    # Método construtor.
    #
    # output_file:
    # Caminho em que o Excel será salvo.
    #
    # score_alert_limit:
    # Limite usado para identificar Scores baixos.
    def __init__(
        self,
        output_file: Path,
        score_alert_limit: int,
    ) -> None:

        # Guarda o caminho do arquivo dentro do objeto.
        self.output_file = output_file

        # Guarda o limite de Score.
        self.score_alert_limit = score_alert_limit

        # Cria um novo arquivo Excel em memória.
        self.workbook = Workbook()

        # Obtém a aba criada automaticamente pelo Workbook.
        self.worksheet: Worksheet = self.workbook.active

        # Altera o nome padrão da aba.
        self.worksheet.title = "Dependências"

        # Chama o método responsável por criar
        # e formatar o cabeçalho.
        self._create_header()

    # Cria o cabeçalho da planilha.
    def _create_header(self) -> None:

        # Converte a tupla HEADERS em lista e adiciona
        # os valores à primeira linha da planilha.
        self.worksheet.append(
            list(self.HEADERS)
        )

        # Percorre todas as células da primeira linha.
        for cell in self.worksheet[1]:

            # Deixa o texto do cabeçalho em negrito.
            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )

            # Define uma cor de fundo azul para o cabeçalho.
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="1F4E78",
            )

            # Centraliza o conteúdo horizontalmente
            # e verticalmente.
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # Define uma altura maior para a primeira linha.
        self.worksheet.row_dimensions[1].height = 24

        # Congela a primeira linha.
        #
        # Ao rolar a planilha para baixo,
        # o cabeçalho continuará visível.
        self.worksheet.freeze_panes = "A2"

    # Adiciona uma dependência à planilha.
    #
    # report deve ser um PackageReportModel,
    # ou seja, o model consolidado pelo Workflow.
    def add(
        self,
        report: PackageReportModel,
    ) -> None:

        # Prepara a data para ser escrita no Excel.
        publication_date = self._prepare_datetime(
            report.last_publication
        )

        # Cria uma lista seguindo exatamente
        # a ordem definida no cabeçalho.
        row = [
            report.name,
            report.project_version or "Não informada",
            report.latest_version or "Não informada",
            report.description or "Não informada",
            report.score,
            report.vulnerabilities,
            report.license or "Não informada",
            publication_date or "Não informada",
        ]

        # Adiciona a lista como uma nova linha.
        self.worksheet.append(row)

        # Obtém o número da linha que acabou de ser adicionada.
        current_row = self.worksheet.max_row

        # Obtém a célula da última publicação.
        #
        # A coluna 8 corresponde à coluna H.
        publication_cell = self.worksheet.cell(
            row=current_row,
            column=8,
        )

        # Se o valor for realmente um datetime,
        # aplica uma formatação visual de data.
        if isinstance(
            publication_cell.value,
            datetime,
        ):
            publication_cell.number_format = (
                "dd/mm/yyyy hh:mm"
            )

        # Centraliza determinadas colunas.
        #
        # 2 → Versão no projeto
        # 3 → Última versão
        # 5 → Score
        # 6 → Vulnerabilidades
        # 8 → Última publicação
        for column_number in (
            2,
            3,
            5,
            6,
            8,
        ):
            self.worksheet.cell(
                row=current_row,
                column=column_number,
            ).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # Permite quebra de linha na descrição.
        self.worksheet.cell(
            row=current_row,
            column=4,
        ).alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

        # Verifica se o Score está disponível
        # e abaixo do limite configurado.
        if (
            report.score is not None
            and report.score
            < self.score_alert_limit
        ):
            # Destaca visualmente toda a linha.
            self._highlight_alert_row(
                row_number=current_row
            )

    # Destaca uma linha cujo Score está abaixo do limite.
    def _highlight_alert_row(
        self,
        row_number: int,
    ) -> None:

        # Cria uma cor de fundo vermelho-claro.
        alert_fill = PatternFill(
            fill_type="solid",
            fgColor="FFC7CE",
        )

        # Cria uma fonte vermelho-escura.
        alert_font = Font(
            color="9C0006",
        )

        # Percorre todas as células da linha.
        for cell in self.worksheet[row_number]:

            # Aplica a cor de fundo.
            cell.fill = alert_fill

            # Aplica a cor da fonte.
            cell.font = alert_font

        # Obtém especificamente a célula do Score.
        score_cell = self.worksheet.cell(
            row=row_number,
            column=5,
        )

        # Deixa o Score em negrito.
        score_cell.font = Font(
            bold=True,
            color="9C0006",
        )

    # Prepara uma data para ser salva no Excel.
    @staticmethod
    def _prepare_datetime(
        value: datetime | None,
    ) -> datetime | None:

        # Se nenhuma data foi recebida,
        # devolve None.
        if value is None:
            return None

        # Verifica se a data possui informação de fuso horário.
        if value.tzinfo is not None:

            # Converte a data para UTC.
            #
            # Depois remove o tzinfo, porque o Excel
            # não armazena fusos horários diretamente.
            return value.astimezone(
                timezone.utc
            ).replace(
                tzinfo=None
            )

        # Se a data já não possui fuso,
        # devolve sem alterações.
        return value

    # Ajusta automaticamente a largura das colunas.
    def _adjust_column_widths(self) -> None:

        # Percorre cada coluna da planilha.
        for column_cells in self.worksheet.columns:

            # Guarda o maior tamanho encontrado
            # naquela coluna.
            maximum_length = 0

            # Obtém a letra da coluna.
            #
            # Por exemplo:
            # 1 → A
            # 2 → B
            column_letter = get_column_letter(
                column_cells[0].column
            )

            # Percorre todas as células da coluna.
            for cell in column_cells:

                # Caso a célula esteja vazia,
                # usa uma string vazia.
                if cell.value is None:
                    cell_text = ""
                else:
                    # Converte o conteúdo para texto.
                    cell_text = str(cell.value)

                # Atualiza o maior tamanho encontrado.
                maximum_length = max(
                    maximum_length,
                    len(cell_text),
                )

            # Adiciona um pequeno espaço extra.
            adjusted_width = maximum_length + 2

            # Impede que uma coluna fique estreita demais.
            adjusted_width = max(
                adjusted_width,
                12,
            )

            # Impede que descrições grandes deixem
            # a planilha excessivamente larga.
            adjusted_width = min(
                adjusted_width,
                60,
            )

            # Aplica a largura calculada.
            self.worksheet.column_dimensions[
                column_letter
            ].width = adjusted_width

    # Salva a planilha no caminho configurado.
    def save(self) -> None:

        # Garante que a pasta output exista.
        #
        # parents=True cria também pastas intermediárias.
        # exist_ok=True evita erro se a pasta já existir.
        self.output_file.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        # Ajusta as larguras antes do salvamento.
        self._adjust_column_widths()

        # Ativa o filtro automático em toda a área
        # preenchida da planilha.
        self.worksheet.auto_filter.ref = (
            self.worksheet.dimensions
        )

        try:
            # Salva o arquivo no disco.
            self.workbook.save(
                self.output_file
            )

        # Trata especificamente o caso comum
        # em que o arquivo está aberto no Excel.
        except PermissionError as error:
            raise SpreadsheetError(
                "Não foi possível salvar a planilha. "
                "Verifique se o arquivo está aberto no Excel: "
                f"{self.output_file}"
            ) from error

        # Trata outros erros de acesso ao sistema de arquivos.
        except OSError as error:
            raise SpreadsheetError(
                "Ocorreu um erro ao salvar a planilha em: "
                f"{self.output_file}"
            ) from error